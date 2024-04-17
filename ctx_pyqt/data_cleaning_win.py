import datetime
import itertools
import os
import sys
import time
import openpyxl
import pandas as pd
import ctx_button_style
import global_var
import pdfplumber
from itertools import product
from PyQt5.Qt import QContextMenuEvent, QPixmap
from PyQt5.QtCore import Qt, pyqtSignal, QThread, QPoint
from PyQt5.QtWidgets import QWidget, QApplication, QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout, \
    QPushButton, QLabel, QLineEdit, QTableWidgetSelectionRange, QMenu, QAction, QComboBox, QFileDialog, QMessageBox, \
    QFileSystemModel, QTreeView, QDockWidget, QTextEdit, QMainWindow, QSpacerItem, QSizePolicy, QAbstractItemView
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from tqdm import tqdm
from Data_clean.ComboCheckBox import ComboCheckBox
from Data_clean.listWidgets import ImageLabel, Usebutton
from Data_clean.stackedWidget import StackedWidget
from search_win import SearchWin
from sub_win import SubWin

from PyPDF2 import PdfFileWriter, PdfReader, PdfWriter
# from paddleocr import PaddleOCR
import logging


# from pdf2docx import Converter
import PyPDF2

class TableWidget(QWidget):
    control_signal = pyqtSignal(list)

    def __init__(self, sheet, type_flag):
        super(TableWidget, self).__init__()
        self.sheet = sheet
        self.type_flag = type_flag
        # self.type_flag：区分sheet和data_list
        self.clipboard = QApplication.clipboard()
        self.__init_ui()

    def clearSelection(self):
        self.table.clearSelection()

    def setRangeSelected(self, res):
        self.table.setRangeSelected(QTableWidgetSelectionRange(res[0], res[1], res[0], res[1]), True)
        item = self.table.item(res[0], res[1])
        self.table.scrollToItem(item, QAbstractItemView.EnsureVisible)

    def __init_ui(self):
        self.table = QTableWidget(100, 20)
        if self.type_flag == 1:
            column_count = self.sheet.max_column
        else:
            column_count = max([len(i) for i in self.sheet])
        self.table.setColumnCount(column_count)
        if self.type_flag == 1:
            show_data = []
            count = 0
            for i in self.sheet.iter_rows(values_only=True):
                if count < 100:
                    show_data.append(i)
                    count += 1
                else:
                    break
            for i in range(100):
                for j in range(column_count):
                    item = QTableWidgetItem(str(show_data[i][j]))
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.table.setItem(i, j, item)
        else:
            for i in range(100):
                for j in range(column_count):
                    item = QTableWidgetItem(str(self.sheet[i][j]))
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.table.setItem(i, j, item)
        self.table.resizeColumnsToContents()
        self.__layout = QVBoxLayout()
        self.__layout.addWidget(self.table)
        self.setLayout(self.__layout)

    def setPageController(self, page):
        """自定义页码控制器"""
        control_layout = QHBoxLayout()
        homePage = QPushButton("首页")
        prePage = QPushButton("<上一页")
        self.curPage = QLabel("1")
        nextPage = QPushButton("下一页>")
        finalPage = QPushButton("尾页")
        self.totalPage = QLabel("共" + str(page) + "页")
        skipLable_0 = QLabel("跳到")
        self.skipPage = QLineEdit()
        skipLabel_1 = QLabel("页")
        confirmSkip = QPushButton("确定")
        homePage.clicked.connect(self.__home_page)
        prePage.clicked.connect(self.__pre_page)
        nextPage.clicked.connect(self.__next_page)
        finalPage.clicked.connect(self.__final_page)
        confirmSkip.clicked.connect(self.__confirm_skip)
        control_layout.addStretch(1)
        control_layout.addWidget(homePage)
        control_layout.addWidget(prePage)
        control_layout.addWidget(self.curPage)
        control_layout.addWidget(nextPage)
        control_layout.addWidget(finalPage)
        control_layout.addWidget(self.totalPage)
        control_layout.addWidget(skipLable_0)
        control_layout.addWidget(self.skipPage)
        control_layout.addWidget(skipLabel_1)
        control_layout.addWidget(confirmSkip)
        control_layout.addStretch(1)
        self.__layout.addLayout(control_layout)

    def update_table(self):
        self.table.clearContents()
        self.current_page = int(self.curPage.text())
        self.start_index = int((self.current_page - 1) * 100)
        if self.type_flag == 1:
            data = [list(row) for row in
                    itertools.islice(self.sheet.iter_rows(), self.start_index, self.start_index + 100)]

            def get_value(x):
                return x.value

            show_data = [[get_value(x) for x in row] for row in data]
            column_count = len(show_data[0])
            for i in range(100):
                for j in range(column_count):
                    # 处理数据不足的情况（最后一页）
                    try:
                        item = QTableWidgetItem(str(show_data[i][j]))
                        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                        self.table.setItem(i, j, item)
                    except:
                        pass
        else:
            self.show_data = self.sheet[self.start_index:int(self.current_page * 100)]
            row_count = len(self.show_data)
            column_count = max([len(i) for i in self.show_data])
            for i in range(row_count):
                for j in range(column_count):
                    item = QTableWidgetItem(str(self.show_data[i][j]))
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.table.setItem(i, j, item)

    def __home_page(self):
        """点击首页信号"""
        self.control_signal.emit(["home", self.curPage.text()])
        self.update_table()

    def __pre_page(self):
        """点击上一页信号"""
        self.control_signal.emit(["pre", self.curPage.text()])
        self.update_table()

    def __next_page(self):
        """点击下一页信号"""
        self.control_signal.emit(["next", self.curPage.text()])
        self.update_table()

    def __final_page(self):
        """尾页点击信号"""
        self.control_signal.emit(["final", self.curPage.text()])
        self.update_table()

    def __confirm_skip(self):
        """跳转页码确定"""
        self.control_signal.emit(["confirm", self.skipPage.text()])
        self.update_table()

    def showTotalPage(self):
        """返回当前总页数"""
        return int(self.totalPage.text()[1:-1])

    def table_copy(self) -> None:
        selectRect = self.table.selectedRanges()
        top, left, bottom, right = None, None, None, None  # 选择区域的边界
        for item in selectRect:  # 获取范围边界
            top = item.topRow()
            left = item.leftColumn()
            bottom = item.bottomRow()
            right = item.rightColumn()
        column_n = right - left + 1
        row_n = bottom - top + 1
        number = row_n * column_n
        copy_text = []
        for i in range(number):
            copy_text.append(' \t')  # 注意，是空格+\t
            if (i % column_n) == (column_n - 1):
                copy_text.append('\n')
            else:
                pass
        copy_text.pop()  # 删去最后多余的换行符
        range1 = range(top, bottom + 1)
        range2 = range(left, right + 1)
        for row, column in product(range1, range2):
            try:
                data = self.table.item(row, column).text()
                number2 = (row - top) * (column_n + 1) + (column - left)
                copy_text[number2] = data + '\t'
                # 计算出单元格的位置，替换掉原来的空格。
            except:
                pass
        self.STR = str()
        for s in copy_text:
            self.STR = self.STR + s
        self.clipboard.setText(self.STR)
        self.STR = str()  # 字符串归零

    def table_paste(self):
        try:  # 有时会误触ctrl+v，避免报错，所以就try了
            row_index = self.index_row
            column_index = self.index_column
            content = self.table.item(row_index, column_index).text()
            paste_text = str()
            for cell in content:
                if cell != '\n':
                    if cell != '\t':
                        paste_text = paste_text + cell
                    else:
                        item = QTableWidgetItem(paste_text)
                        self.table.setItem(row_index, column_index, item)
                        paste_text = ''
                        column_index += 1
                else:
                    item = QTableWidgetItem(paste_text)
                    self.table.setItem(row_index, column_index, item)
                    paste_text = ''
                    row_index += 1
                    column_index = self.index_column
                item = QTableWidgetItem(paste_text)
                self.table.setItem(row_index, column_index, item)
        except:
            pass

    def table_cut(self) -> None:
        selectRect = self.table.selectedRanges()
        top, left, bottom, right = None, None, None, None  # 选择区域的边界
        for item in selectRect:  # 获取范围边界
            top = item.topRow()
            left = item.leftColumn()
            bottom = item.bottomRow()
            right = item.rightColumn()
        column_n = right - left + 1
        row_n = bottom - top + 1
        number = row_n * column_n
        copy_text = []
        for i in range(number):
            copy_text.append(' \t')  # 注意，是空格+\t
            if (i % column_n) == (column_n - 1):
                copy_text.append('\n')
            else:
                pass
        copy_text.pop()  # 删去最后多余的换行符
        range1 = range(top, bottom + 1)
        range2 = range(left, right + 1)
        for row, column in product(range1, range2):
            try:
                data = self.table.item(row, column).text()
                number2 = (row - top) * (column_n + 1) + (column - left)
                copy_text[number2] = data + '\t'
                # 计算出单元格的位置，替换掉原来的空格。
            except:
                pass
        self.STR = str()
        for s in copy_text:
            self.STR = self.STR + s
        self.clipboard.setText(self.STR)
        self.STR = str()  # 字符串归零
        # 填充为空
        range1 = range(top, bottom + 1)
        range2 = range(left, right + 1)
        for i in range1:
            for j in range2:
                fill = ""
                item = QTableWidgetItem(fill)
                self.table.setItem(i, j, item)

    def select_all(self):
        row_count = len(self.data_list)
        # 取列最长那个
        column_count = max([len(i) for i in self.data_list])
        self.table.setRangeSelected(QTableWidgetSelectionRange(0, 0, row_count - 1, column_count - 1),
                                    True)  # 起始行，起始列，结束行，结束列

    def contextMenuEvent(self, evt: QContextMenuEvent) -> None:
        # 打印一条消息到控制台，表示右键菜单被触发
        # 创建一个QMenu对象，该对象将包含我们的菜单项
        right_menu = QMenu(self)
        # right_undo = QAction("撤销", right_menu)
        right_cut = QAction("剪切", right_menu)
        right_copy = QAction("复制", right_menu)
        right_paste = QAction("粘贴", right_menu)
        right_clear = QAction("清除内容", right_menu)
        right_insert = QMenu("插入单元格", self)
        cell_right = QAction("活动单元格右移", self)
        cell_below = QAction("活动单元格下移", self)
        cell_rows = QAction("整行", self)
        cell_cols = QAction("整列", self)
        right_insert.addActions([cell_right, cell_below, cell_rows, cell_cols])
        cell_del = QMenu("删除单元格", right_menu)
        right_cell_left = QAction("右侧单元格左移", self)
        below_cell_up = QAction("下方单元格上移", self)
        cell_rows_del = QAction("整行", self)
        cell_cols_del = QAction("整列", self)
        cell_del.addActions([right_cell_left, below_cell_up, cell_rows_del, cell_cols_del])
        cell_sort = QAction("排序", right_menu)
        cell_set_format = QAction("设置单元格格式", right_menu)
        # right_menu.addAction(right_undo)
        right_menu.addAction(right_cut)
        right_menu.addAction(right_copy)
        right_menu.addAction(right_paste)
        right_menu.addAction(right_clear)
        right_menu.addAction(cell_sort)
        right_menu.addAction(cell_set_format)
        right_cut.triggered.connect(self.table_cut)
        right_copy.triggered.connect(self.table_copy)
        right_paste.triggered.connect(self.table_paste)
        right_cut.triggered.connect(self.table_cut)
        # right_undo.setShortcut("U")
        # right_cut.setShortcut("T")
        # right_copy.setShortcut("C")
        # right_paste.setShortcut("P")
        # right_clear.setShortcut("N")
        # cell_sort.setShortcut("O")
        # cell_set_format.setShortcut("F")
        right_menu.addMenu(right_insert)
        right_menu.addMenu(cell_del)
        # 显示菜单并等待用户选择一个选项，然后关闭菜单。用户选择的选项将返回给调用者
        # point 调用
        right_menu.exec_(evt.globalPos())


class DataInfo(QWidget):
    def __init__(self, str1):
        super().__init__()
        self.resize(300, 100)
        self.resize(300, 100)
        main_layout = QVBoxLayout()

        # 文本框
        label = QLabel("")
        str2 = str1
        label.setText(str2)
        """文件共有几行，几列
        文件有多少缺失值，哪列存在多少多少缺失值
        文件共有多少重复值，哪列存在多少重复值
        文件各个字段类型"""
        main_layout.addWidget(label)
        self.setLayout(main_layout)


class PdfMenu(QWidget):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("pdf文件处理")
        layout = QHBoxLayout()
        pdf_split_button = ButtonImage("pdf文件分割", 'icons/文件.png')
        pdf_text_button = ButtonImage("pdf提取文本", 'icons/EXCEL.png')
        pdf_table_button = ButtonImage("pdf提取表格", 'icons/pdf.png')
        pdf_ocr_button = ButtonImage("图片提取文字", 'icons/EXCEL.png')
        pdf_docx_button = ButtonImage("pdf转word", 'icons/pdf.png')
        pdf_split_button.clicked.connect(self.split_pdf)
        pdf_text_button.clicked.connect(self.extract_text)
        pdf_table_button.clicked.connect(self.extract_excel)
        pdf_ocr_button.clicked.connect(self.pdf_ocr)
        pdf_docx_button.clicked.connect(self.pdf_to_docx)
        layout.addWidget(pdf_split_button)
        layout.addWidget(pdf_text_button)
        layout.addWidget(pdf_table_button)
        layout.addWidget(pdf_ocr_button)
        layout.addWidget(pdf_docx_button)

        for item in (pdf_split_button, pdf_text_button, pdf_table_button, pdf_ocr_button, pdf_docx_button):
            item.setFixedSize(100, 100)
            item.setStyleSheet(ctx_button_style.button_style())
        self.setLayout(layout)

        # 将窗体移动到屏幕中心
        # self.move(center_x, center_y)

    def clear_dir(self, dir_path):
        """清空目录下的文件"""
        names = os.listdir(dir_path)
        for name in names:
            file_path = os.path.join(dir_path, name)
            cmd = 'del %s' % file_path  # 这个是windows命令
            # cmd = 'rm -rf %s' % file_path  #这个是Linux的命令
            cmd = cmd.replace('/', '\\')  # 为啥要这个？因为windows命令不支持/所以需要替换，Linux命令没有遇见这个问题；
            os.system(cmd)  # 可以通过返回值来判断是否执行成功：0成功，其他失败

    def split_pdf(self):
        """pdf按每页拆分"""
        # 获取 PdfFileReader 对象
        default_path = os.getcwd()
        dlg = QFileDialog(None, "choose_file", default_path, 'All Files(*)')
        dlg.setFileMode(QFileDialog.AnyFile)
        if dlg.exec_():
            pdf_file = dlg.selectedFiles()[0]
        reader = PdfReader(pdf_file)
        output_path = os.path.dirname(pdf_file)
        pages_num = len(reader.pages)
        # writer = PdfFileWriter() 生成一个文件
        for index in range(pages_num):
            # 可以通过对index判断分割想要的
            writer = PdfWriter()  # 按照每页来分割pdf
            pageObj = reader.pages[index]
            writer.add_page(pageObj)
            # 添加完每页，再一起保存至文件中；如果要输出一个文件，后面这些放置到循环外即可
            file_name = os.path.join(output_path, str(index) + '.pdf')
            with open(file_name, 'wb') as fw:
                writer.write(fw)
        msg_box = QMessageBox(QMessageBox.Information, '完成提示', '分割文件完成！')
        msg_box.exec_()

    def extract_text(self):
        default_path = os.getcwd()
        dlg = QFileDialog(None, "choose_file", default_path, 'All Files(*)')
        dlg.setFileMode(QFileDialog.AnyFile)
        if dlg.exec_():
            pdf = dlg.selectedFiles()[0]
        output_path = os.path.dirname(pdf)
        file_name = os.path.splitext(pdf)[0] + ".txt"
        wookroot = pdfplumber.open(pdf)
        with open(os.path.join(output_path, file_name), 'w', encoding='utf-8') as file:
            pages = wookroot.pages
            for page in pages:
                text = page.extract_text()
                file.write(text + '\n')  # 在每个文本后面添加换行符，以便于区分不同页面的文本

        msg_box = QMessageBox(QMessageBox.Information, '完成提示', '提取文本完成！')
        msg_box.exec_()

    def extract_excel(self):
        default_path = os.getcwd()
        dlg = QFileDialog(None, "choose_file", default_path, 'All Files(*)')
        dlg.setFileMode(QFileDialog.AnyFile)
        if dlg.exec_():
            pdf = dlg.selectedFiles()[0]
        output_path = os.path.dirname(pdf)
        file_name = os.path.splitext(pdf)[0] + ".xlsx"
        wookroot = pdfplumber.open(pdf)

        pages = wookroot.pages
        for page in pages:
            tables = page.extract_tables()
        # 创建Excel工作簿和工作表
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # 将表格写入Excel工作表中
        for i, table in enumerate(tables):
            for j, row in enumerate(table):
                for k, cell in enumerate(row):
                    worksheet.cell(row=j + 1, column=k + 1, value=cell)

                    # 保存Excel文件
        workbook.save(os.path.join(output_path, file_name))

        msg_box = QMessageBox(QMessageBox.Information, '完成提示', '提取表格完成！')
        msg_box.exec_()

    def pdf_ocr(self):
        # 将一些日志信息过滤掉

        # 创建文件选择对话框
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.ExistingFiles)  # 设置文件模式为现有文件

        # 设置过滤器，只显示图片文件
        dialog.setNameFilter("Images (*.png *.xpm *.jpg)")

        # 显示对话框并获取选择的文件列表
        if dialog.exec_() == QFileDialog.Accepted:
            img = dialog.selectedFiles()[0]  # 获取选择的文件列表

        logging.disable(logging.DEBUG)
        logging.disable(logging.WARNING)

        ocr = PaddleOCR(use_angle_cls=True, lang='ch')
        result = ocr.ocr(img, cls=True)
        if len(result) != 1:
            print('result num=', len(result))
            exit()
        result = result[0]
        lines = []
        output_path = os.path.dirname(img)
        file_name = os.path.splitext(img)[0] + ".txt"
        # 可以直接答应result看看都是些什么结构
        for line in result:
            lines.append(line[1][0])
        # 这就是识别出来的文本
        with open(os.path.join(output_path, file_name), "w") as file:
            for item in lines:
                file.write("%s\n" % item)
        msg_box = QMessageBox(QMessageBox.Information, '完成提示', '图片提取文字完成完成！')
        msg_box.exec_()

    def pdf_to_docx(self):

        default_path = os.getcwd()
        dlg = QFileDialog(None, "choose_file", default_path, 'All Files(*)')
        dlg.setFileMode(QFileDialog.AnyFile)
        if dlg.exec_():
            pdf = dlg.selectedFiles()[0]
        output_path = os.path.dirname(pdf)
        file_name = os.path.splitext(pdf)[0] + ".docx"
        logging.disable(logging.INFO)
        logging.disable(logging.DEBUG)
        logging.disable(logging.WARNING)
        cv = Converter(pdf)
        cv.convert(os.path.join(output_path, file_name))
        cv.close()
        msg_box = QMessageBox(QMessageBox.Information, '完成提示', 'pdf转word完成！')
        msg_box.exec_()


class ZScore(QWidget):
    signal_zscore = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.resize(300, 100)
        main_layout = QVBoxLayout()
        layout_z = QHBoxLayout()
        layout_m = QHBoxLayout()
        label_max_min = QLabel("选择标准化方法")
        self.comb_m = QComboBox()
        self.comb_m.addItem("Z-Score标准化")
        self.comb_m.addItem("Max-Min标准化")
        layout_m.addWidget(label_max_min)
        layout_m.addWidget(self.comb_m)
        label = QLabel("选择要进行z-score的字段")
        col_name = global_var.get_value("col_name")
        self.comb = ComboCheckBox(col_name)
        ok_button = QPushButton("确定！")
        ok_button.clicked.connect(self.ok_event)
        layout_z.addWidget(label)
        layout_z.addWidget(self.comb)
        main_layout.addLayout(layout_m)
        main_layout.addLayout(layout_z)
        main_layout.addWidget(ok_button)
        self.setLayout(main_layout)

    def ok_event(self):
        res_list = self.comb.currentText()
        res_list = [item for item in res_list.split(";") if item != ""]
        res_list.append(self.comb_m.currentText())
        self.signal_zscore.emit(res_list)


class FileMerge(QWidget):

    def __init__(self, parentwin=None):
        super().__init__()
        self.setWindowTitle("多文件合并")
        layout = QHBoxLayout()
        txt_button = ButtonImage("文本文件", 'icons/文件.png')
        excel_button = ButtonImage("Excel文件", 'icons/EXCEL.png')
        pdf_button = ButtonImage("pdf文件", 'icons/pdf.png')
        txt_button.clicked.connect(self.text_file)
        excel_button.clicked.connect(self.excel_file)
        pdf_button.clicked.connect(self.pdf_file)
        layout.addWidget(txt_button)
        layout.addWidget(excel_button)
        layout.addWidget(pdf_button)
        for item in (txt_button, excel_button, pdf_button):
            item.setFixedSize(100, 100)
            item.setStyleSheet(ctx_button_style.button_style())
        self.setLayout(layout)
        globalPos = parentwin.mapToGlobal(QPoint(0, 0))
        # 计算新位置的 x 和 y 坐标
        x = globalPos.x() + (parentwin.width() - self.width()) // 2
        y = globalPos.y() + (parentwin.height() - self.height()) // 2
        # 移动窗口到新位置
        self.move(x, y)
        # 将窗体移动到屏幕中心
        # self.move(center_x, center_y)

    def text_file(self):
        dir_choose = QFileDialog.getExistingDirectory(self,
                                                      "选取文件夹",
                                                      os.getcwd())  # 起始路径
        file_list = os.listdir(dir_choose)
        # 遍历文件列表，过滤出文件并排除文件夹
        file_list = [f for f in file_list if os.path.isfile(os.path.join(dir_choose, f))]
        file_result_name = str(datetime.datetime.now())[:-7].replace("-", "").replace(":", "").replace(" ",
                                                                                                       "") + "merge" + \
                           os.path.splitext(file_list[0])[1]
        with open(os.path.join(dir_choose, file_result_name),
                  'w') as out:
            for file in file_list:
                with open(os.path.join(dir_choose, file), 'r') as f:
                    out.writelines(f.readlines())
                    out.writelines("\n")
        msg_box = QMessageBox(QMessageBox.Information, '完成提示', '合并文件完成！')
        msg_box.exec_()

    def excel_file(self):
        dir_choose = QFileDialog.getExistingDirectory(self,
                                                      "选取文件夹",
                                                      os.getcwd())  # 起始路径
        file_list = os.listdir(dir_choose)
        frames = []
        # 遍历文件列表，过滤出文件并排除文件夹
        file_list = [f for f in file_list if os.path.isfile(os.path.join(dir_choose, f))]
        file_result_name = str(datetime.datetime.now())[:-7].replace("-", "").replace(":", "").replace(" ",
                                                                                                       "") + "merge" + \
                           os.path.splitext(file_list[0])[1]
        for file in file_list:
            df = pd.read_excel(os.path.join(dir_choose, file))  # excel转换成DataFrame
            frames.append(df)
        result = pd.concat(frames)

        result.to_excel(os.path.join(dir_choose, file_result_name), index=False)
        msg_box = QMessageBox(QMessageBox.Information, '完成提示', '合并文件完成！')
        msg_box.exec_()

    def pdf_file(self):
        dir_choose = QFileDialog.getExistingDirectory(self,
                                                      "选取文件夹",
                                                      os.getcwd())  # 起始路径
        pdf_writer = PyPDF2.PdfWriter()

        for filename in os.listdir(dir_choose):
            if filename.endswith('.pdf'):
                pdf_reader = PyPDF2.PdfReader(os.path.join(dir_choose, filename))
                for page_num in range(len(pdf_reader.pages)):
                    pdf_writer.add_page(pdf_reader.pages[page_num])
        file_list = os.listdir(dir_choose)
        file_result_name = str(datetime.datetime.now())[:-7].replace("-", "").replace(":", "").replace(" ",
                                                                                                       "") + "merge" + \
                           os.path.splitext(file_list[0])[1]
        with open(os.path.join(dir_choose, file_result_name), 'wb') as output_file:
            pdf_writer.write(output_file)
            print("合并完成")
        msg_box = QMessageBox(QMessageBox.Information, '完成提示', '合并文件完成！')
        msg_box.exec_()


class ButtonImage(QPushButton):
    def __init__(self, name_label, image_url):
        super().__init__()
        self.resize(100, 100)
        hbox = QVBoxLayout()
        self.name_label = QLabel(name_label, self)
        self.name_label.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.image_label = QLabel(self)
        self.image_label.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        pixmap = QPixmap(image_url)
        self.image_label.setPixmap(pixmap)
        hbox.addWidget(self.image_label)
        hbox.addWidget(self.name_label)
        self.setLayout(hbox)


class Worker(QThread):
    progress = pyqtSignal(str)
    signal_data_list = pyqtSignal(list)

    def __init__(self, sheet):
        super().__init__()
        self.sheet = sheet

    def run(self):
        self.data_list = []
        self.pbar = tqdm(total=self.sheet.max_row)
        for i in tqdm(self.sheet.iter_rows(values_only=True)):
            self.data_list.append(i)
            self.pbar.update(1)  # 更新进度条
            if len(self.data_list) % 20000 == 0:
                self.progress.emit(str(self.pbar))
        self.signal_data_list.emit(self.data_list)


class Menu(QMainWindow):
    signal_main_to_usebutton = pyqtSignal(list)

    def __init__(self):
        super(Menu, self).__init__()
        self.SubWin = None
        self.res_list = None
        self.search = None
        self.main_text = None
        self.STR = None
        self.as_save_date = None
        self.save_date = None
        self.data_list = None
        self.file_path = None
        self.table_model = None
        self.file_model = None
        self.is_text_edit = None
        self.select_list = None
        self.init_ui()

    def init_ui(self):
        disk_dock = QDockWidget('目录', self)
        disk_dock.setMinimumSize(150, 100)  # 宽度，高度
        disk_dock.setMaximumSize(150, 1600)  # 宽度，高度
        self.file_model = QFileSystemModel()
        self.file_model.setRootPath('.')
        file_tree = QTreeView()
        file_tree.setModel(self.file_model)
        file_tree.setColumnWidth(0, 200)
        file_tree.setColumnHidden(1, True)
        file_tree.setColumnHidden(2, True)
        file_tree.setColumnHidden(3, True)
        file_tree.clicked.connect(self.on_clicked)
        file_tree.header().hide()
        disk_dock.setWidget(file_tree)
        disk_dock.setFloating(False)
        disk_dock.setFeatures(QDockWidget.NoDockWidgetFeatures)  # 设置不可脱离
        self.addDockWidget(Qt.LeftDockWidgetArea, disk_dock)
        self.setStyleSheet("QMainWindow::separator { width: 0px; height: 0px; margin: 0px; padding: 0px; }")
        self.back_win = QWidget(self)
        self.back_win.setObjectName("test")
        self.back_win.setStyleSheet(
            "#test{border-image: url(C:/Users/Administrator/PycharmProjects/deal_tool/CTX/ctx_pyqt/icon/背景1.webp);}")
        self.setCentralWidget(self.back_win)
        self.funcListWidget = ImageLabel(self)
        self.dock_func = QDockWidget("数据清洗", self)
        title_win = QWidget(self)
        title_win.setObjectName("itemwin")
        title_win.setStyleSheet("#itemwin{ background-color: #FFFFFF; }")
        layout = QHBoxLayout()
        search_button = ButtonImage("查找", 'icon/查找.png')
        sub_button = ButtonImage("替换", 'icon/替换.png')
        multi_file_button = ButtonImage("多文件合并", 'icon/多文件合并.png')
        z_score_button = ButtonImage("数据标准化", 'icon/数据标准化.png')
        pdf_button = ButtonImage("PDF文件处理", 'icon/PDF.png')
        data_check_button = ButtonImage("数据概况", 'icon/数据概况.png')
        save_button = ButtonImage("保存", 'icon/保存.png')
        as_save_button = ButtonImage("另存为", 'icon/另存为.png')
        cofig_button = ButtonImage("清洗配置", 'icon/配置文件.png')
        for item in (
                search_button, sub_button, multi_file_button, z_score_button,
                pdf_button,
                data_check_button,save_button,as_save_button,cofig_button):
            item.setFixedSize(100, 100)
            item.setStyleSheet(ctx_button_style.button_style())
        # 添加到布局
        spacer = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        layout.addWidget(search_button, 0, Qt.AlignLeft)
        layout.addWidget(sub_button, 0, Qt.AlignLeft)
        layout.addWidget(multi_file_button, 0, Qt.AlignLeft)
        layout.addWidget(z_score_button, 0, Qt.AlignLeft)
        layout.addWidget(pdf_button, 0, Qt.AlignLeft)
        layout.addWidget(data_check_button, 0, Qt.AlignLeft)
        layout.addWidget(save_button, 0, Qt.AlignLeft)
        layout.addWidget(as_save_button, 0, Qt.AlignLeft)
        layout.addWidget(cofig_button, 0, Qt.AlignLeft)
        layout.addSpacerItem(spacer)
        layout.setContentsMargins(0, 0, 0, 0)
        title_win.setLayout(layout)
        search_button.clicked.connect(self.search_show)
        sub_button.clicked.connect(self.sub_show)
        multi_file_button.clicked.connect(self.show_file_merge)
        pdf_button.clicked.connect(self.show_pdf_mean)
        z_score_button.clicked.connect(self.show_z_score)
        data_check_button.clicked.connect(self.show_data_info)
        as_save_button.clicked.connect(self.as_save)
        self.dock_func.setMinimumSize(100, 150)  # 宽度，高度
        self.dock_func.setMaximumSize(2600, 1600)  # 宽度，高度
        self.dock_func.setTitleBarWidget(title_win)
        self.dock_func.setContentsMargins(0, 0, 0, 0)
        self.dock_func.setWidget(self.funcListWidget)
        self.dock_func.setFeatures(QDockWidget.NoDockWidgetFeatures)
        self.dock_func.setStyleSheet("QDockWidget::title{background : #BFEFFF;}")
        self.dock_func.setDisabled(True)
        self.addDockWidget(Qt.TopDockWidgetArea, self.dock_func)
        self.useListWidget = Usebutton(self)
        self.dock_used = QDockWidget('已选操作', self)
        self.dock_used.setMinimumSize(150, 100)  # 宽度，高度
        self.dock_used.setMaximumSize(600, 1600)  # 宽度，高度
        self.dock_used.setWidget(self.useListWidget)
        self.dock_used.setFeatures(QDockWidget.NoDockWidgetFeatures)
        self.addDockWidget(Qt.RightDockWidgetArea, self.dock_used)
        self.stackedWidget = StackedWidget(self)
        self.dock_attr = QDockWidget('属性', self)
        self.dock_attr.setFixedSize(400, 200)
        self.dock_attr.setWidget(self.stackedWidget)
        self.dock_attr.setFeatures(QDockWidget.NoDockWidgetFeatures)
        self.dock_attr.close()
        self.addDockWidget(Qt.RightDockWidgetArea, self.dock_attr)
        # 状态信息窗口
        self.info_win = QTextEdit(self)
        self.dock_info = QDockWidget('日志信息', self)
        self.dock_info.setWidget(self.info_win)
        self.dock_info.setFeatures(QDockWidget.NoDockWidgetFeatures)
        self.addDockWidget(Qt.BottomDockWidgetArea, self.dock_info)
        self.dock_info.setMinimumSize(100, 100)  # 宽度，高度
        self.dock_info.setMaximumSize(2000, 100)  # 宽度，高度
        self.src_img = None
        self.cur_img = None
        self.setWindowTitle("ADC")

    def startProgress(self):
        self.worker = Worker(self.sheet)
        self.worker.progress.connect(self.updateProgressBar)
        self.worker.signal_data_list.connect(self.updatedatalist)
        self.info_win.append("文件开始加载")
        self.worker.start()

    def read_data_deal(self):
        columns = self.data_list[0]
        self.df = pd.DataFrame(self.data_list[1:], columns=columns)
        rows, cols = self.df.shape
        duplicates = self.df.duplicated()
        num_duplicates = duplicates.sum()
        missing_counts = {col: self.df[col].isnull().sum() for col in self.df.columns}
        total = sum(missing_counts.values())
        self.file_info = "文件共有" + str(rows) + "行，" + str(cols) + "列\n" + "文件共有" + str(
            total) + "个缺失值，其中：\n"
        str1 = ""
        for item in missing_counts.items():
            str1 = str1 + f"{item[0]}列缺失{item[1]}个" + "\n"
        self.file_info = self.file_info + str1 + "文件共有" + str(num_duplicates) + "个缺失值"
        self.list1 = list(self.df.columns)
        self.list2 = list(self.df.columns)
        global_var._init()
        # 所有列名转字符串类型
        self.list1 = [str(item) for item in self.list1]
        print("写入global_var", self.list1)
        global_var.set_value("col_name", self.list1)
        # 文件名，文件类型，文件路径，文件大小，创建时间，修改时间
        self.info_win.append(
            str(datetime.datetime.now())[:-7] + "   " + "文件共计" + str(self.df.shape[0]) + "行," + str(
                self.df.shape[1]) + "列" + "\n")

        self.useListWidget.signal_ctx_satrt.connect(self.deal_emit_start_slot)
        self.row_count = len(self.data_list)
        # 取列最长那个
        self.columnc_ount = max([len(i) for i in self.data_list])

    def updateProgressBar(self, value):
        self.info_win.append(value)

    def updatedatalist(self, res):
        self.info_win.append("文件全部加载完成")
        self.dock_func.setDisabled(False)
        self.data_list = res
        self.read_data_deal()
        self.table_widget = TableWidget(self.data_list, 2)  # 实例化表格
        self.table_widget.setPageController(int(self.sheet.max_row / 100) + 1)  # 表格设置页码控制
        self.table_widget.control_signal.connect(self.page_controller)
        self.setCentralWidget(self.table_widget)

    def timer_decorator(func):
        def wrapper(self, *args, **kwargs):
            start_time = time.time()
            result = func(self, *args, **kwargs)
            end_time = time.time()
            print(f"Function {func.__name__} took {end_time - start_time} seconds to run.")
            return result

        return wrapper

    def show_data_info(self):
        self.data_info = DataInfo(self.file_info)
        self.data_info.show()

    def show_z_score(self):
        self.z = ZScore()
        self.z.show()
        self.z.signal_zscore.connect(self.deal_emit_zscore_slot)

    def deal_emit_zscore_slot(self, res_list):
        X_test = self.df[res_list[:-1]]  # 实例化对象
        if res_list[-1] == "Z-Score标准化":
            b_test = StandardScaler()  # 训练数据，赋值给b_test
        else:
            b_test = MinMaxScaler()
        X_result = b_test.fit_transform(X_test)
        for index, item in enumerate(res_list[:-1]):
            self.df.loc[:, item] = X_result[:, index]

        total_rows = self.df.shape[0]
        column_names = list(self.df.columns)
        self.data_list = []
        self.data_list.insert(0, column_names)
        for i in range(total_rows):
            df_row = self.df.iloc[i].tolist()
            self.data_list.append(df_row)
        self.input_data()

    def show_file_merge(self):
        self.file_merge = FileMerge(self)
        self.file_merge.show()

    def show_pdf_mean(self):
        self.pdf_mean = PdfMenu()
        self.pdf_mean.show()

    def update_image(self):
        if self.src_img is None:
            return
        img = self.process_image()
        self.cur_img = img
        self.graphicsView.update_image(img)

    def change_image(self, img):
        self.src_img = img
        img = self.process_image()
        self.cur_img = img
        self.graphicsView.change_image(img)

    def process_image(self):
        img = self.src_img.copy()
        for i in range(self.useListWidget.count()):
            img = self.useListWidget.item(i)(img)
        return img

    def on_clicked(self, index):
        file_path = self.file_model.filePath(index)
        if file_path:
            self.file_path = file_path
            if self.file_path.split(".")[-1] in ("xlsx", "xls", "txt", "csv"):
                self.read()
            else:
                msg_box = QMessageBox(QMessageBox.Critical, '错误提示', '请选择xlsx，xls，txt，csv中的一种格式')
                msg_box.exec_()

    def selected(self, index):  # 获取选中单元格的索引
        self.index_row = index.row()
        self.index_column = index.column()

    def handleCellChanged(self):
        self.select_list = []
        sele = self.table.selectedIndexes()
        for i in sele:
            self.select_list.append((i.row(), i.column()))
        row_list = []
        column_list = []
        for item in self.select_list:
            if item[0] not in row_list:
                row_list.append(item[0])
            if item[1] not in column_list:
                column_list.append(item[1])

    def page_controller(self, signal):
        total_page = self.table_widget.showTotalPage()
        if "home" == signal[0]:
            self.table_widget.curPage.setText("1")
        elif "pre" == signal[0]:
            if 1 == int(signal[1]):
                QMessageBox.information(self, "提示", "已经是第一页了", QMessageBox.Yes)
                return
            self.table_widget.curPage.setText(str(int(signal[1]) - 1))
        elif "next" == signal[0]:
            if total_page == int(signal[1]):
                QMessageBox.information(self, "提示", "已经是最后一页了", QMessageBox.Yes)
                return
            self.table_widget.curPage.setText(str(int(signal[1]) + 1))
        elif "final" == signal[0]:
            self.table_widget.curPage.setText(str(total_page))
        elif "confirm" == signal[0]:
            if total_page < int(signal[1]) or int(signal[1]) < 0:
                QMessageBox.information(self, "提示", "跳转页码超出范围", QMessageBox.Yes)
                return
            self.table_widget.curPage.setText(signal[1])

        self.changeTableContent()  # 改变表格内容

    def changeTableContent(self):
        """根据当前页改变表格的内容"""
        cur_page = self.table_widget.curPage.text()

    @timer_decorator
    def open(self):
        """选择文件对话框"""
        # QFileDialog组件定义
        file_dialog = QFileDialog(self)
        # QFileDialog组件设置
        file_dialog.setWindowTitle("选择文件")  # 设置对话框标题
        file_dialog.setFileMode(QFileDialog.AnyFile)  # 设置能打开文件的格式
        file_dialog.setDirectory(r'C:\Users\Administrator\Desktop\CTX')  # 设置默认打开路径
        file_dialog.setNameFilter("数据源(*.csv *.xlsx)")  # 按文件名过滤
        self.file_path = file_dialog.exec()  # 窗口显示，返回文件路径
        if self.file_path and file_dialog.selectedFiles():
            self.file_path = file_dialog.selectedFiles()[0]
            self.read()

    @timer_decorator
    def read(self):
        self.input_flag = 1
        file_end = self.file_path.split(".")[-1]  # 文件类型
        if file_end == "xlsx":
            workbook = openpyxl.load_workbook(filename=self.file_path, read_only=True)
            self.sheet = workbook.active
            self.startProgress()
        elif file_end == "csv":
            # 直接显示csv
            self.df = pd.read_csv(self.file_path)
            df_list = self.df.values.tolist()
            headers = self.df.columns.tolist()
            self.data_list = [headers] + df_list
            self.read_data_deal()
        else:
            self.df = None
            msg_box = QMessageBox(QMessageBox.Critical, '错误提示', '请选择xlsx, csv中的一种格式')
            msg_box.exec_()
        self.input_data()

    @timer_decorator
    def input_data(self):
        """写入数据"""
        if self.file_path.split(".")[-1] == "xlsx":
            if self.input_flag == 1:
                self.table_widget = TableWidget(self.sheet, 1)  # 实例化表格
                self.table_widget.setPageController(int(self.sheet.max_row / 100) + 1)  # 表格设置页码控制
                self.table_widget.control_signal.connect(self.page_controller)
                self.setCentralWidget(self.table_widget)
                self.input_flag = 2
            else:
                self.table_widget = TableWidget(self.data_list, 2)  # 实例化表格
                self.table_widget.setPageController(int(self.sheet.max_row / 100) + 1)  # 表格设置页码控制
                self.table_widget.control_signal.connect(self.page_controller)
                self.setCentralWidget(self.table_widget)
        else:
            self.table_widget = TableWidget(self.data_list, 2)  # 实例化表格
            self.dock_func.setDisabled(False)
            self.table_widget.setPageController(int(len(self.data_list) / 100) + 1)  # 表格设置页码控制
            self.table_widget.control_signal.connect(self.page_controller)
            self.setCentralWidget(self.table_widget)

    def save_file(self):
        # 读取表格数据，再存储到原路径
        row_count = self.table.model().rowCount()
        column_count = self.table.model().columnCount()
        # 用二维列表存储
        self.save_date = []
        for i in range(row_count):
            line = []
            for j in range(column_count):
                item = self.table.model().item(i, j)
                text = item.text()
                line.append(text)
            self.save_date.append(line)
        file_end = self.file_path.split(".")[-1]
        if file_end in ("xlsx", "xls"):
            df = pd.DataFrame(self.save_date)
            df = df.reset_index(drop=True)
            df.to_excel(self.file_path, index=False, header=None)
        elif file_end == "csv":
            df = pd.DataFrame(self.save_date)
            df = df.reset_index(drop=True)
            df.to_csv(self.file_path, index=False, header=None)
        elif file_end == "txt":
            df = pd.DataFrame(self.save_date)
            df = df.reset_index(drop=True)
            df.to_csv(self.file_path, sep='\t', index=False, header=None)

    def new_file(self):
        pass

    def as_save(self):
        """文件另存对话框"""
        # 显示文件另存对话框
        options = '文本文档(*.txt);;Excel 工作簿(*.xlsx);;CSV (逗号分隔)(*.csv)'
        file_path, _ = QFileDialog.getSaveFileName(self, '文件保存', '', options)

        if file_path:
            # 获取文件扩展名
            file_end = file_path.split(".")[-1].lower()

            # 根据文件扩展名决定保存格式
            if file_end == 'txt':
                self.save_as_txt(file_path)
            elif file_end in ['xlsx', 'xls']:
                self.save_as_excel(file_path)
            elif file_end == 'csv':
                self.save_as_csv(file_path)
            else:
                print("不支持的文件格式")
                # 读取表格数据，再存储到原路径

    def save_as_txt(self, file_path):
        """保存为文本文件"""
        with open(file_path, 'w', encoding='utf-8') as f:
            for row in self.data_list:
                f.write('\t'.join(str(item) for item in row) + '\n')
        print(f"数据已保存为文本文件到 {file_path}")

    def save_as_csv(self, file_path):
        """保存为CSV文件"""
        df = pd.DataFrame(self.data_list)
        df.to_csv(file_path, index=False, encoding='utf-8-sig')
        print(f"数据已保存为CSV文件到 {file_path}")

    def save_as_excel(self, file_path):
        """保存为Excel文件"""
        df = pd.DataFrame(self.data_list)
        df.to_excel(file_path, index=False, engine='openpyxl')
        print(f"数据已保存为Excel文件到 {file_path}")

    @timer_decorator
    def deal_emit_start_slot(self, res_dict):
        # 缺失值处理
        # 考虑各个选项的互斥关系
        # {'NanTableWidget': {'填充为指定值': '231', '删除空值所在的行': '是', '删除空值所在的列': '是', '填充为其他': '前一个观测值'}}
        try:
            if res_dict['NanTableWidget']['删除空值所在的行'] == "是":
                # 删除包含空值的行
                self.df = self.df.dropna()
            if res_dict['NanTableWidget']['删除空值所在的列'] == "是":
                # 删除包含空值的列
                self.df = self.df.dropna(axis=1)
            if res_dict['NanTableWidget']['填充为指定值'] != "":
                self.df = self.df.fillna(value=res_dict['NanTableWidget']['填充为指定值'])
            if res_dict['NanTableWidget']['填充为其他'] not in ["", "否"]:
                if res_dict['NanTableWidget']['填充为其他'] == "前一个观测值":
                    self.df = self.df.fillna(method='ffill')
                if res_dict['NanTableWidget']['填充为其他'] == "后一个观测值":
                    self.df = self.df.fillna(method='bfill')
                if res_dict['NanTableWidget']['填充为其他'] == "平均值":
                    self.df = self.df.fillna(self.df.mean())
                if res_dict['NanTableWidget']['填充为其他'] == "中位数":
                    self.df = self.df.fillna(self.df.median())
                if res_dict['NanTableWidget']['填充为其他'] == "众数":
                    # 分别对每列使用众数填充缺失值
                    for column in self.df.columns:
                        mode = self.df[column].mode()[0]  # 获取众数
                        self.df[column].fillna(mode, inplace=True)
                self.df = self.df.fillna(value=res_dict['NanTableWidget']['填充为其他'])
        except:
            pass
        # 重复值处理
        try:
            if res_dict['RepeatTabledWidget']['删除重复行'] == "所有列重复":
                self.df = self.df.drop_duplicates()
            if res_dict['RepeatTabledWidget']['删除重复行'][:-2] in self.list1:
                self.df = self.df.drop_duplicates(res_dict['RepeatTabledWidget']['删除重复行'][:-2])
            if res_dict['RepeatTabledWidget']['删除重复列'] == "是":
                self.df = self.df.T.drop_duplicates().T
        except:
            pass
        # 重命名列名
        try:
            if res_dict['ColTabledWidget']['列名重命名'] != "":
                new_col_name = res_dict['ColTabledWidget']['列名重命名'].split(",")
                self.df.columns = new_col_name
        except:
            pass
        # 类型转换
        try:
            type_dict = res_dict['FormatTabledWidget']
            type_to_fun = {"小数类型": 'float64', '整数类型': 'int64', '文本': 'str'}
            for item in self.list2:
                if type_dict[item] != "不修改":
                    if type_dict[item] != "日期时间":
                        self.df[item] = self.df[item].astype(eval(type_to_fun[type_dict[item]]))
                    else:
                        self.df[item] = pd.to_datetime(self.df[item])
        except:
            pass
        """'IlocTableWidget': {'': False, '大于': '11', '介于下限': '22', '介于上限': '33', '小于': '2112', '文本包含内容': 'bnbm,', '数值范围选择字段': 'A列;B列;C列;', 
        '文本包含字段': 'C列;D列;'}}"""
        # 数据筛选
        # 大于
        for item in res_dict['IlocTableWidget']['数值范围选择字段'].split(';'):
            try:
                self.df = self.df[self.df[item] > float(res_dict['IlocTableWidget']['大于'])]
            except:
                pass
        # 小于
        for item in res_dict['IlocTableWidget']['数值范围选择字段'].split(';'):
            try:
                self.df = self.df[self.df[item] < float(res_dict['IlocTableWidget']['小于'])]
            except:
                pass
        # 介于
        for item in res_dict['IlocTableWidget']['数值范围选择字段'].split(';'):
            try:
                self.df = self.df[(self.df[item] > float(res_dict['IlocTableWidget']['介于下限'])) & (
                        self.df[item] < float(res_dict['IlocTableWidget']['介于上限']))]
            except:
                pass
        # 文本包含
        for item in res_dict['IlocTableWidget']['文本包含字段'].split(';'):
            try:
                self.df = self.df[self.df[item].str.contains(res_dict['IlocTableWidget']['文本包含内容'])]
            except:
                pass
        # 数据排序
        # 'DatasortTableWidget': {'A列': '升序', 'B列': '升序', 'C列': '升序', 'D列': '不排序', 'E列': '不排序', 'F列': '不排序'}
        for item in res_dict['DatasortTableWidget'].keys():
            if res_dict['DatasortTableWidget'][item] != "不排序":
                try:
                    self.df = self.df.sort_values(by=item, ascending=True if res_dict['DatasortTableWidget'][
                                                                                 item] == "升序" else False)
                except KeyError:
                    # 处理整数列名的情况
                    self.df = self.df.sort_values(by=int(item), ascending=True if res_dict['DatasortTableWidget'][
                                                                                      item] == "升序" else False)
        # 异常值处理
        # 'DataoverTableWidget': {'A列': '填为平均值', 'B列': '填为平均值', 'C列': '不处理', 'D列': '不处理', 'E列': '填为平均值', 'F列': '删除该行'}
        for item in res_dict['DataoverTableWidget'].keys():
            if res_dict['DataoverTableWidget'][item] != "不处理":
                mean = self.df[item].mean()
                std_dev = self.df[item].std()
                lower_bound = mean - 3 * std_dev
                upper_bound = mean + 3 * std_dev
                # ['不处理', '删除该行', '填为0', '填为平均值', '填为中位数', '填为众数']
                if res_dict['DataoverTableWidget'][item] == "删除该行":
                    self.df.drop(self.df[(self.df[item] < lower_bound) | (self.df[item] > upper_bound)].index,
                                 inplace=True)
                if res_dict['DataoverTableWidget'][item] == "填为0":
                    self.df.loc[(self.df[item] < lower_bound) | (self.df[item] > upper_bound), item] = 0
                if res_dict['DataoverTableWidget'][item] == "填为平均值":
                    self.df[(self.df[item] < lower_bound) | (self.df[item] > upper_bound)] = mean
                if res_dict['DataoverTableWidget'][item] == "填为中位数":
                    median = self.df[item].median()
                    self.df[(self.df[item] < lower_bound) | (self.df[item] > upper_bound)] = median
                if res_dict['DataoverTableWidget'][item] == "填为众数":
                    mode = self.df['数值'].mode()[0]  # 获取众数，可能存在多个众数，因此需要使用索引[0]获取第一个众数
                    self.df[(self.df[item] < lower_bound) | (self.df[item] > upper_bound)] = mode
        print("信号接收成功", res_dict)
        total_rows = self.df.shape[0]
        column_names = list(self.df.columns)
        self.data_list = []
        self.data_list.insert(0, column_names)
        for i in range(total_rows):
            df_row = self.df.iloc[i].tolist()
            self.data_list.append(df_row)
        self.input_data()

    def search_show(self):
        self.main_text = self.data_list
        # 根据内容查找索引，再选中
        self.search = SearchWin(self.main_text)
        self.search.signal_def.connect(self.deal_emit_search_slot)
        self.search.show()

    def sub_show(self):
        self.main_text = self.data_list
        self.SubWin = SubWin(self.main_text)
        self.SubWin.signal_def.connect(self.deal_emit_search_slot)
        self.SubWin.signal_sub_def.connect(self.deal_emit_sub_slot)
        self.SubWin.signal_sub_all_def.connect(self.deal_emit_sub_all_slot)
        self.SubWin.show()

    def deal_emit_search_slot(self, res):
        self.table_widget.clearSelection()
        self.table_widget.setRangeSelected(res)

    def deal_emit_sub_slot(self, res):
        """
        替换当前
        :param res:匹配到的当前结果索引
        :return:
        """
        if self.is_text_edit:
            start_index = res[0][0]  # 坐标点索引行坐标
            end_index = res[0][1]  # 坐标点索引列坐标
            res_index = res[1]  # 坐标点列表
            res_count = res[2]  # 当前坐标点在坐标点列表中的索引
            sub_text = res[3]  # 替换内容
            tmp_text = self.main_text[:start_index] + sub_text + self.main_text[end_index:]
            self.main_text_edit.setText(tmp_text)
            self.main_text = self.main_text_edit.toPlainText()
            try:
                self.deal_emit_search_slot(res_index[res_count + 1])  # 选中下一个,最后一个会越界
            except:
                pass
        else:
            row_index = res[0][0]
            column_index = res[0][1]
            res_index = res[1]  # 坐标点列表
            res_count = res[2]  # 当前坐标点在坐标点列表中的索引
            sub_text = res[3]  # 替换内容
            old_text = res[4]
            tmp_text = self.table_widget.table.item(row_index, column_index).text().replace(old_text, sub_text)
            # 如果不区分大小写，大写小写文本都要替换成新文本
            item = QTableWidgetItem(tmp_text)
            self.table_widget.table.setItem(row_index, column_index, item)
            try:
                self.deal_emit_search_slot(res_index[res_count + 1])  # 选中下一个,最后一个会越界
            except:
                pass

    def deal_emit_sub_all_slot(self, res):
        """
        替换所有
        :param res:匹配到的单元格索引
        :return:
        """
        if self.is_text_edit:
            old_text = res[0]
            sub_text = res[1]
            tmp_text = self.main_text.replace(old_text, sub_text)
            self.main_text_edit.setText(tmp_text)
            self.main_text = self.main_text_edit.toPlainText()
        else:
            old_text = res[0]
            sub_text = res[1]
            res_index = res[2]  # 坐标点列表
            for index_tuple in res_index:
                tmp_text = self.table_widget.table.item(index_tuple[0], index_tuple[1]).text().replace(old_text,
                                                                                                       sub_text)
                item = QTableWidgetItem(tmp_text)
                self.table_widget.setItem(index_tuple[0], index_tuple[1], item)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    LW = Menu()
    LW.show()
    sys.exit(app.exec_())
